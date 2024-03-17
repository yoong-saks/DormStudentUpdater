import tkinter as tk
import openpyxl
from tkinter import filedialog
import tkinter.messagebox as messagebox
from datetime import datetime, timedelta
from tkinter import ttk
import pyperclip
import os
import win32com.client as win32
import shutil

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.main_excel_file = ''
        self.sub_excel_file = ''
        self.date = datetime.today().strftime('%Y/%m/%d')
        self.create_widgets()

    def create_widgets(self):
        self.master.title("사생명단 자동 업데이트 푸로그램 by 정호성")

        self.windows_usr_name = os.path.expanduser('~')
        backup_dir_path = f"{self.windows_usr_name}\\Desktop\\자동엑셀백업"
        
        if not os.path.exists(backup_dir_path):
            os.mkdir(backup_dir_path)

        self.main_excel_label = tk.Label(self.master, text="공유폴더 엑셀파일: " + self.main_excel_file)
        self.main_excel_label.pack()

        self.filepathname = '.path_autoexcel'
        self.filepathnamepath = os.path.join(backup_dir_path, self.filepathname)
        if os.path.exists(self.filepathnamepath):
            with open(self.filepathnamepath, "r") as f:
                self.last_selected_path = f.read()
                self.main_excel_file = self.last_selected_path
                self.main_excel_label.config(text="공유폴더 엑셀파일: " + self.last_selected_path)

        self.main_excel_button = tk.Button(self.master, text="Browse", command=self.browse_main_excel)
        self.main_excel_button.pack()

        self.sub_excel_label = tk.Label(self.master, text="사생명단 엑셀파일: " + self.sub_excel_file)
        self.sub_excel_label.pack()

        self.sub_excel_button = tk.Button(self.master, text="Browse", command=self.browse_sub_excel)
        self.sub_excel_button.pack()

        self.start_button = tk.Button(self.master, text="프로그램 시작", command=self.start_program)
        self.start_button.pack()

    def browse_main_excel(self):
        self.main_excel_file = filedialog.askopenfilename()
        self.main_excel_label.config(text="공유폴더 엑셀파일: " + self.main_excel_file)

        with open(self.filepathnamepath, "w") as f:
            f.write(self.main_excel_file)

        

    def browse_sub_excel(self):
        self.sub_excel_file = filedialog.askopenfilename()
        self.sub_excel_label.config(text="사생명단 엑셀파일: " + self.sub_excel_file)

        
    def start_program(self):

        def xls_to_xlsx(file_path):
            if file_path.endswith('.xls'):
                now_dir = os.path.abspath('./')
                
                excel = win32.Dispatch('Excel.Application')
                wb = excel.Workbooks.Open(file_path)
                new_file_path = os.path.splitext(file_path)[0] + '.xlsx'
                new_file_path = new_file_path.replace('/', '\\')
                
                wb.SaveAs(new_file_path, FileFormat=51)
                wb.Close()
                excel.Application.Quit()
                os.remove(file_path)
                return new_file_path
            else:
                return file_path

        selected_date_obj = datetime.strptime(self.date, '%Y/%m/%d')
       
        month = selected_date_obj.month
        day = selected_date_obj.day

        backup_dir_path = f"{self.windows_usr_name}\\Desktop\\자동엑셀백업"
        backup_name = f'{month}월 {day}일 사생명단 백업.xlsx'
        copy_file_path = os.path.join(backup_dir_path, backup_name)

        if os.path.exists(copy_file_path):
            suffix = 1
            while True:
                new_file_name = f"{backup_name[:-5]} ({suffix}).xlsx"
                new_thisfile_path = os.path.join(backup_dir_path, new_file_name)
                if os.path.exists(new_thisfile_path):
                    suffix += 1
                else:
                    copy_file_path = new_thisfile_path
                    break

        shutil.copy(self.main_excel_file,copy_file_path)
        


        self.wb_main = openpyxl.load_workbook(self.main_excel_file)
        self.ws_main = self.wb_main['사생명단']

        converted_sub_excel_file = xls_to_xlsx(self.sub_excel_file)
        self.wb_sub = openpyxl.load_workbook(converted_sub_excel_file)
        self.ws_sub = self.wb_sub.active


        for row in self.ws_main.iter_rows():
            for cell in row:
                if cell.value == '학번':
                    row_index = cell.row
                    self.ws_main.delete_rows(row_index + 1, self.ws_main.max_row - row_index)


        for row in self.ws_sub.iter_rows():
            for cell in row:
                if cell.value == 'L R':
                    subrow_index = cell.row
                    subcol_index = cell.column

        main_row_index = row_index+1
        for row in range(subrow_index+1, self.ws_sub.max_row + 1):
            if self.ws_sub.cell(row=row, column=subcol_index).value:
                self.ws_main.cell(row=main_row_index, column=1).value = self.ws_sub.cell(row=row, column=1).value
                self.ws_main.cell(row=main_row_index, column=2).value = self.ws_sub.cell(row=row, column=2).value
                self.ws_main.cell(row=main_row_index, column=3).value = self.ws_sub.cell(row=row, column=3).value
                self.ws_main.cell(row=main_row_index, column=4).value = self.ws_sub.cell(row=row, column=4).value
                self.ws_main.cell(row=main_row_index, column=5).value = self.ws_sub.cell(row=row+1, column=1).value
                main_row_index += 1
        

        self.wb_main.save(self.main_excel_file)

        messagebox.showinfo("진행 완료 됐따~~", "사생명단이 업데이트 됐다굿!")

root = tk.Tk()
app = Application(master=root)
app.mainloop()
